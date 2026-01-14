import io
import tempfile
from datetime import date

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st

# Assume modules are in the same folder
import FCST_M1 as M1
import FCST_M2 as M2
import FCST_M3 as M3


COLOR_MAP = {
    "SELLOUT": "#ff00ea",       # slateblue
    "ESHIP": "#1f77b4",         # default blue
    "ASHIP": "#ff7f0e",         # orange
    "Previous_FCST": "#2ca02c", # green
    "Proposed_FCST": "#d62728", # red
    "Proposed_Preview": "#d62728",
}


# ─────────────────────────
# (Fixed) trend parameters
# ─────────────────────────
TREND_LOOKBACK_FIXED = 3
TREND_RAMP_FIXED     = 4
TREND_UP_FIXED       = 0.15
TREND_DOWN_FIXED     = 0.15
M3.TREND_CAP_ABS        = 0.60
M3.SINGLE_SPIKE_WEIGHT  = 0.85
M3.SINGLE_SPIKE_FACTOR  = 2.0
M3.SINGLE_SPIKE_TOL     = 0.05

# ─────────────────────────
# Session state initialization
# ─────────────────────────
if "review_results" not in st.session_state:
    st.session_state.review_results = None
if "ran_review" not in st.session_state:
    st.session_state.ran_review = False

# Manual adjustment store {(Material, ds) -> pct (decimal)}
if "fcst_adj" not in st.session_state:
    st.session_state.fcst_adj = {}

# ★★★ Dict for preview (immediate graph reflection)
if "fcst_adj_preview" not in st.session_state:
    st.session_state.fcst_adj_preview = {}

# ★★★ Manual toggle override store for Update Needed {(Material, ds) -> bool}
if "update_needed_overrides" not in st.session_state:
    st.session_state.update_needed_overrides = {}


# ★ Buffer preview & apply store (global/by-month)
if "buf_global_pct" not in st.session_state:
    st.session_state.buf_global_pct = {}  # {mat: float (e.g., 0.10)}

if "buf_month_preview" not in st.session_state:
    st.session_state.buf_month_preview = {}  # {(mat, ds_ts): float}

# (Optional) Track applied buffer details if desired
if "buf_month_applied" not in st.session_state:
    st.session_state.buf_month_applied = {}  # {(mat, ds_ts): float}


# -----------------------------
# Shared utilities
# -----------------------------


# === Segment standardization utilities ===
import unicodedata
import re as _re

SEGMENT_CANON_MAP = {
    "impress": "imPRESS",
    "preglued nails": "PreGlued Nails",
    "french nails": "French Nails",
    "decorated nails": "Decorated Nails",
    "color nails": "Color Nails",
    "toe nails": "Toe Nails",
    "impress toe nail": "Toe Nails",
}

def _canon_seg_series(s: pd.Series) -> pd.Series:
    return (s.astype(str)
              .apply(lambda x: unicodedata.normalize("NFKC", x))   # Unicode normalization
              .str.replace(r"\s+", " ", regex=True)                # Collapse multiple spaces → single space
              .str.strip()
              .str.lower())

def normalize_segment(series: pd.Series) -> pd.Series:
    key = _canon_seg_series(series)
    return key.replace(SEGMENT_CANON_MAP)

def canon_key_series(s: pd.Series) -> pd.Series:
    out = s.astype(str).str.strip()
    return out.str.replace(r"^0+(?!$)", "", regex=True)

def normalize_yearmonth(col: pd.Series) -> pd.Series:
    s = pd.to_datetime(col, errors="coerce")
    need_fix = s.isna()
    if need_fix.any():
        s2 = pd.to_datetime(col.astype(str), format="%Y%m", errors="coerce")
        s = s.mask(need_fix, s2)
    return s.dt.to_period("M").dt.to_timestamp(how="start")

def period_start(ts) -> pd.Timestamp:
    return pd.to_datetime(ts).to_period("M").to_timestamp(how="start")

def save_uploader_to_temp(upl, suffix=".xlsx") -> str:
    if upl is None:
        return ""
    data = upl.read()
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tf.write(data)
    tf.flush()
    tf.close()
    return tf.name

def df_to_excel_download(df: pd.DataFrame, filename: str, label: str):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    st.download_button(
        label=label,
        data=buf.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ==== Report Export Helper (Updated Needed only; Old vs New by Month) ====
def build_update_report_df(final_fcst: pd.DataFrame, cutoff_ts: pd.Timestamp) -> pd.DataFrame:
    df = final_fcst.copy()
    df["ds"] = pd.to_datetime(df["ds"], errors="coerce")
    df["Month"] = df["ds"].dt.to_period("M").astype(str)

    # After cutoff + update-needed only
    h_start = cutoff_ts + pd.offsets.MonthBegin(1)
    mask = (df["ds"] >= h_start) & (df["FCST_Update_Needed"].fillna(False))
    sub = df.loc[mask, ["MATERIAL_KEY", "Month", "previous_FCST", "FCST_final"]].copy()
    if sub.empty:
        # Keep table shape even if empty
        return pd.DataFrame(columns=["ITEM"])

    # Sort months
    months = sorted(sub["Month"].unique(), key=lambda s: pd.Period(s))

    # Old / New pivot
    prev_piv = (
        sub.pivot_table(index="MATERIAL_KEY", columns="Month", values="previous_FCST", aggfunc="sum")
          .reindex(columns=months)
    )
    new_piv = (
        sub.pivot_table(index="MATERIAL_KEY", columns="Month", values="FCST_final", aggfunc="sum")
          .reindex(columns=months)
    )
    prev_piv = np.rint(prev_piv).astype("Int64")
    new_piv  = np.rint(new_piv).astype("Int64")

    # Two-row header: (Old/New, Month)
    prev_piv.columns = pd.MultiIndex.from_product([["Old"], prev_piv.columns])
    new_piv.columns  = pd.MultiIndex.from_product([["New"], new_piv.columns])

    report = pd.concat([prev_piv, new_piv], axis=1)
    report = report.reset_index().rename(columns={"MATERIAL_KEY": "ITEM"})

    report["ITEM"] = report["ITEM"].astype(str)
    return report

# ==== Final FCST export helpers ====

def _mark_needs_update(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in ["source", "Group", "FCST_final", "USW_baseline", "M3_baseline", "ASHIP_4M"]:
        if c not in df.columns:
            df[c] = np.nan
    needs = df["source"].ne("KEEP")

    baseline_any = pd.Series(False, index=df.index)
    for bc in ["USW_baseline", "M3_baseline", "ASHIP_4M"]:
        if bc in df.columns:
            baseline_any = baseline_any | df[bc].notna()

    needs |= df["FCST_final"].isna() & baseline_any
    return df.assign(needs_update=needs.astype(bool))

def _normalize_final_schema(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    REQUIRED_ORDER = [
        "Retailer", "Brand", "Material", "Desc",
        "Segment", "Group", "source", "LongShort",
        "FCST_previous", "FCST_final",
        "USW_baseline", "ASHIP_4M", "trend_multiplier",
        "note"
    ]
    for c in REQUIRED_ORDER:
        if c not in df.columns:
            df[c] = np.nan
    id_like = ["Retailer","Brand","Material","Desc","Segment","Group","source","LongShort","note"]
    num_like = ["FCST_previous","FCST_final","USW_baseline","ASHIP_4M","trend_multiplier"]
    for c in id_like:
        if c in df.columns:
            df[c] = df[c].astype("string")
    for c in num_like:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    ordered = [c for c in REQUIRED_ORDER if c in df.columns]
    rest = [c for c in df.columns if c not in ordered]
    df = df[ordered + rest]
    if "Retailer" in df.columns and "Material" in df.columns:
        df = df.sort_values(["Retailer","Material"])
    return df

def _to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")

def safe_take(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    keep = [c for c in cols if c in df.columns]
    if not keep:
        return pd.DataFrame()
    return df[keep]

# ─────────────────────────
# Progress/log utilities
# ─────────────────────────
def make_progress_widgets(place="main"):
    container = st.sidebar if place == "sidebar" else st
    bar = container.progress(0, text="Ready")
    status = container.status("FCST Review running…", expanded=True)

    def update(pct: float, msg: str):
        pct_i = int(max(0, min(100, pct)))
        bar.progress(pct_i, text=msg)
        with status:
            st.write(f"{pct_i}% · {msg}")

    def complete(msg: str = "Complete!"):
        bar.progress(100, text=msg)
        status.update(label=msg, state="complete")

    return update, complete

# -----------------------------
# Snowflake loader (with close handling)
# -----------------------------
def make_snowflake_loader(user, account, warehouse, database, schema, authenticator="externalbrowser"):
    import snowflake.connector
    def _loader(brand: str):
        conn = snowflake.connector.connect(
            user=user, account=account, authenticator=authenticator,
            warehouse=warehouse, database=database, schema=schema
        )
        cur = None
        try:
            cur = conn.cursor()
            cur.execute(f'USE WAREHOUSE {warehouse}')
            cur.execute(f'USE DATABASE {database}')
            cur.execute(f'USE SCHEMA {schema}')
            customer_key = M1.CUSTOMER_MAP.get(brand.upper())
            if not customer_key:
                raise ValueError(f"Unknown brand: {brand} (CUSTOMER_MAP에 없음)")
            q = f"""
            SELECT "MATERIAL_KEY", "MONTH_KEY", "ESHIP", "ASHIP+Open" AS "ASHIP"
            FROM ZPPRFC01
            WHERE "PLANT_KEY" IN ('G100') AND "CUSTOMER_KEY" = '{customer_key}'
            """
            df = pd.read_sql(q, conn)
            df.rename(columns={"MONTH_KEY": "YearMonth"}, inplace=True)
            df["YearMonth"] = normalize_yearmonth(df["YearMonth"].astype(str))
            if "MATERIAL_KEY" in df.columns:
                df["MATERIAL_KEY"] = canon_key_series(df["MATERIAL_KEY"])
            return df
        finally:
            try:
                if cur is not None:
                    cur.close()
            finally:
                conn.close()
    return _loader

# (Optional) ASHIP only
def make_snowflake_loader_aship_only(user, account, warehouse, database, schema, authenticator="externalbrowser"):
    import snowflake.connector
    def _loader(brand: str):
        conn = snowflake.connector.connect(
            user=user, account=account, authenticator=authenticator,
            warehouse=warehouse, database=database, schema=schema
        )
        cur = None
        try:
            cur = conn.cursor()
            cur.execute(f'USE WAREHOUSE {warehouse}')
            cur.execute(f'USE DATABASE {database}')
            cur.execute(f'USE SCHEMA {schema}')
            customer_key = M1.CUSTOMER_MAP.get(brand.upper())
            if not customer_key:
                raise ValueError(f"Unknown brand: {brand} (CUSTOMER_MAP에 없음)")
            q = f"""
            SELECT "MATERIAL_KEY", "MONTH_KEY", "ASHIP"
            FROM ZPPRFC01
            WHERE "PLANT_KEY" IN ('G100') AND "CUSTOMER_KEY" = '{customer_key}'
            """
            df = pd.read_sql(q, conn)
            df.rename(columns={"MONTH_KEY": "YearMonth"}, inplace=True)
            df["YearMonth"] = normalize_yearmonth(df["YearMonth"].astype(str))
            df["MATERIAL_KEY"] = canon_key_series(df["MATERIAL_KEY"])
            return df[["MATERIAL_KEY","YearMonth","ASHIP"]]
        finally:
            try:
                if cur is not None:
                    cur.close()
            finally:
                conn.close()
    return _loader

# -----------------------------
# Sidebar: (simplified) Use Trend + Flag thresholds
# -----------------------------

def render_review_sidebar():
    st.sidebar.header("⚙️ Review Settings")
    use_trend = st.sidebar.checkbox("Use Trend Adjustment", value=True)

    # 🔹 Select review comparison window (3M / 4M)
    compare_months = st.sidebar.radio("Review window (months)", [3, 4], index=1)

    st.sidebar.markdown("☣️ Gap Flag Thresholds")
    thr1 = st.sidebar.number_input("Threshold for POS - ESHIP", 0.0, 1.0, 0.15, 0.01, format="%.2f")
    thr2 = st.sidebar.number_input("Threshold for POS - ASHIP", 0.0, 1.0, 0.15, 0.01, format="%.2f")
    thr3 = st.sidebar.number_input("Threshold for ESHIP - ASHIP", 0.0, 1.0, 0.15, 0.01, format="%.2f")

    # 🔎 Further Review Needed (keep computation; remove filter)
    fr_thr = st.sidebar.number_input("Further Review: total diff ratio thr", 0.0, 2.0, 0.30, 0.01, format="%.2f")
    fr_months = st.sidebar.number_input("Further Review: horizon (months)", 1, 12, 6, 1)

    # 🔹 GrossChange (sum-based % change) threshold
    st.sidebar.markdown("☣️ Gross Change Threshold")
    gross_thr = st.sidebar.number_input("Threshold (sum %)", 0.0, 1.0, 0.20, 0.01, format="%.2f")

    return (
        use_trend,
        float(thr1), float(thr2), float(thr3),
        float(gross_thr),
        int(compare_months),
        float(fr_thr), int(fr_months)
    )




# -----------------------------
# Main UI (keep existing screen)
# -----------------------------
st.title("FCST App - by MinCho's Cotton Candy")

st.subheader("Ganppare Mincho, Ganppare JJ")
mode = st.radio("Mode", ["Normal FCST", "FCST Review"], horizontal=True)
st.divider()

# Sidebar default-value guard
use_trend = False


flag_thr1 = flag_thr2 = flag_thr3 = 0.30
gross_thr = 0.20

if mode == "FCST Review":
    use_trend, flag_thr1, flag_thr2, flag_thr3, gross_thr, compare_months, fr_thr, fr_months = render_review_sidebar()


# Hide sidebar in Normal mode (optional)
if mode == "Normal FCST":
    st.markdown("""
    <style>
      [data-testid="stSidebar"] {display: none;}
      .block-container {padding-left: 2rem; padding-right: 2rem;}
    </style>
    """, unsafe_allow_html=True)

# Form (clear_on_submit=False) — keep existing UI
with st.form("main_form", clear_on_submit=False):
    st.markdown("### File Upload & Snowflake Connection")
    c1, c2, c3 = st.columns([1,1,1])
    with c1:
        brand = st.selectbox(
            "Customer (Brand)",
            list(M1.CUSTOMER_MAP.keys()),
            index=list(M1.CUSTOMER_MAP.keys()).index("WG") if "WG" in M1.CUSTOMER_MAP else 0
        )
        pos_file = st.file_uploader("POS File(.xlsx)", type=["xlsx"])
        usw_file = st.file_uploader("USW File(.xlsx)", type=["xlsx"])
    with c2:
        sf_user  = st.text_input("Snowflake User", value="JEJEON@KISSUSA.COM")
        sf_acct  = st.text_input("Snowflake Account", value="UKDVSEA-NPB82638")
        sf_wh    = st.text_input("Warehouse", value="COMPUTE_WH")
        sf_db    = st.text_input("Database", value="KDB")
        sf_schema= st.text_input("Schema", value="SCP")
        sf_auth  = st.text_input("Authenticator", value="externalbrowser")
    
    st.markdown("### FCST Parameters")
    c3, c4, c5 = st.columns(3)
    with c3:
        cutoff = st.date_input("CUTOFF (1st Day of Month)", value=date(2025, 9, 1))
        periods = st.number_input("PERIODS (Number of Months)", min_value=1, max_value=24, value=6, step=1)
    with c4:
        weeks_pm = st.number_input("WEEKS_PM (Month-Week Conversion)", min_value=3.5, max_value=5.0, value=4.3, step=0.1)
        min_hist = st.number_input("MIN_HISTORY (Long/Short)", min_value=6, max_value=48, value=24, step=1)
    with c5:
        band_low  = st.number_input("BAND LOW", min_value=0.5, max_value=1.0, value=0.89, step=0.01, format="%.6f")
        band_high = st.number_input("BAND HIGH", min_value=1.0, max_value=1.5, value=1.21, step=0.01, format="%.6f")

    run_btn = st.form_submit_button("🚀 RUN")

# Stop before RUN, but keep showing previous run results if available
if (not run_btn) and (not st.session_state.ran_review):
    st.stop()

# On RUN: save uploads & inject parameters
if run_btn:
    pos_path = save_uploader_to_temp(pos_file) if pos_file else ""
    usw_path = save_uploader_to_temp(usw_file) if usw_file else ""

    if mode == "Normal FCST":
        if not pos_path:
            st.error("Upload POS file."); st.stop()
        if not usw_path:
            st.error("Upload USW file. (Normal FCST mode)"); st.stop()
    elif mode == "FCST Review":
        if not pos_path:
            st.error("Upload POS file."); st.stop()
        if not usw_path:
            st.warning("If USW file is not uploaded, baseline may be 0 in USW-based forecasting stage (M2).")

    # Inject module-level global parameters
    M1.WEEKS_PM = weeks_pm
    M1.CUTOFF   = period_start(cutoff)
    for m in (M2, M3):
        m.WEEKS_PM    = weeks_pm
        m.CUTOFF      = period_start(cutoff)
        m.PERIODS     = int(periods)
        m.MIN_HISTORY = int(min_hist)
        m.BAND_LOW, m.BAND_HIGH = float(band_low), float(band_high)

cutoff_ts = period_start(cutoff)

# -------------------------------------------------
# Execute
# -------------------------------------------------
if mode == "Normal FCST":
    if not run_btn:
        st.stop()
    st.subheader("🟦 Normal FCST (Full SKU FCST using U/S/W)")
    try:
        # ✅ Always run via M2: Door@cutoff > 0 filter and Long/Short seasonality handled inside M2
        out_raw, info = M2.run_usw_fcst(
            pos_path=pos_path,
            usw_path=usw_path,
            brand=brand,
            cutoff=cutoff_ts,
            periods=int(periods),
            # Keep defaults for drop_* / precomputed, etc.
        )

        # ✅ Normalize to the UI standard schema
        out = out_raw.copy()
        if "Month" not in out.columns and "ds" in out.columns:
            out["Month"] = pd.to_datetime(out["ds"]).dt.to_period("M").astype(str)

        rename_map = {"ds": "YearMonth", "factor": "Seasonality", "yhat_denorm": "FCST"}
        keep = [c for c in ["Material", "ds", "Month", "method", "factor", "baseline_units", "yhat_denorm"]
                if c in out.columns]
        out = out[keep].rename(columns=rename_map)

        st.success("Success!")
        if isinstance(info, dict):
            c1, c2, c3 = st.columns(3)
            c1.metric("CUTOFF", info.get("cutoff", f"{cutoff_ts:%Y-%m-01}"))
            c2.metric("periods", str(info.get("periods", int(periods))))
            c3.metric("brand", info.get("brand", brand))

        st.subheader("Results Preview (U/S/W based FCST)")
        st.dataframe(out, use_container_width=True, height=480)

        csv_buf = io.StringIO()
        out.to_csv(csv_buf, index=False)
        st.download_button(
            "📥 Download Full Results CSV",
            data=csv_buf.getvalue().encode("utf-8"),
            file_name=f"{brand}_FCST_{cutoff_ts:%Y%m}.csv",
            mime="text/csv",
            use_container_width=True
        )
        df_to_excel_download(out, f"Normal_FCST_{brand}_{cutoff_ts:%Y%m}.xlsx", "⬇️ Output Download (XLSX)")
    except Exception as e:
        st.exception(e)


else:
    # ─────────────────────────────────────────────────────────
    # 🟧 FCST Review
    # ─────────────────────────────────────────────────────────
    st.subheader("🟧 FCST Review")

    # Create progress widgets
    progress_cb, progress_done = make_progress_widgets(place="main")

    # One-time Snowflake fetch + cache
    @st.cache_data(show_spinner=False)
    def sf_fetch_all(user, account, warehouse, database, schema, authenticator, brand):
        loader = make_snowflake_loader(user, account, warehouse, database, schema, authenticator)
        return loader(brand)
    


    def compute_further_review_flags(
        review_df: pd.DataFrame,
        final_fcst: pd.DataFrame,
        sf_fcst: pd.DataFrame,
        cutoff: pd.Timestamp,
        horizon_months: int = 6,
        diff_ratio_thr: float = 0.30,
        model_col: str = "FCST_final",
        sf_col: str = "Snowflake_FCST",
        require_full: bool = True,          # ← Added: default keeps existing behavior (evaluate only full months)
        min_coverage_ratio: float = 0.75,   # ← Added: minimum coverage for partial allowance (e.g., ≥3 of 4 months)
        target_scope: str = "ALL"
    ) -> pd.DataFrame:
        """
        KEEP(그룹 '000' 또는 source='KEEP') 대상. cutoff 포함 horizon 개월 합계 기준으로
        모델 vs Snowflake 총합 차이비율(|ΣM-ΣS|/ΣS)이 임계 초과면 FurtherReviewNeeded=1.
        단 Snowflake의 월 데이터가 horizon 개월 모두 존재할 때만 평가.
        """
        if review_df is None or review_df.empty:
            return review_df

        # Months to compare (including cutoff)
        months = pd.period_range(
            cutoff.to_period("M"),
            (cutoff.to_period("M") + (horizon_months - 1)),
            freq="M"
        ).to_timestamp(how="start")

        # Normalize
        def _norm(df):
            df = df.copy()
            # 1) Choose source date column: prefer YearMonth, else use ds
            if "YearMonth" in df.columns:
                ym = df["YearMonth"]
            elif "ds" in df.columns:
                ym = df["ds"]
            else:
                raise KeyError("Expected a 'YearMonth' or 'ds' column for monthly alignment.")

            # 2) Safe parsing: datetime → normalize to month start (supports YYYYMM/strings)
            ym_dt = pd.to_datetime(ym, errors="coerce")

            if ym_dt.isna().any():
                # Handle YYYYMM or mixed strings
                ym_str = ym.astype(str).str.replace(r"[^0-9]", "", regex=True)
                ym_fallback = pd.to_datetime(ym_str + "01", format="%Y%m%d", errors="coerce")
                ym_dt = ym_dt.fillna(ym_fallback)

            df["YearMonth"] = ym_dt.dt.to_period("M").dt.to_timestamp(how="start")
            return df


        ffc = _norm(final_fcst)
        sfc = _norm(sf_fcst)

        # Slice to the period
        ffc = ffc[ffc["YearMonth"].isin(months)]
        sfc = sfc[sfc["YearMonth"].isin(months)]

        # Sum (SKU-level horizon total)
# Compare only SKU×month intersection (safe even when partial is allowed)
# 1) Cut model/Snowflake to horizon, then inner-join on intersecting months
        aligned = (
            ffc[["MATERIAL_KEY","YearMonth",model_col]]
            .merge(
                sfc[["MATERIAL_KEY","YearMonth",sf_col]],
                on=["MATERIAL_KEY","YearMonth"], how="inner"
            )
        )

        # 2) Coverage calculation (number of intersecting months per SKU)
        coverage = (aligned.groupby("MATERIAL_KEY")["YearMonth"]
                        .nunique()
                        .rename("Covered_Months"))

        # 3) Sum (based on the intersection)
        agg = (aligned.groupby("MATERIAL_KEY", as_index=True)
                    .agg(Model_Sum=(model_col,"sum"),
                        SF_Sum=(sf_col,"sum")))

        # 4) Determine completeness / partial-allowance status
        # 4) Determine completeness / partial-allowance status
        target_months = len(months)
        coverage_ratio = (coverage / target_months).rename("Coverage_Ratio")
        complete_sf = (coverage == target_months)

        # Determine which SKUs are eligible for evaluation
        if require_full:
            eval_ok = complete_sf
        else:
            eval_ok = (coverage_ratio >= float(min_coverage_ratio))

        # Build into a single table
        meta = pd.concat([coverage, coverage_ratio, complete_sf.rename("SF_Complete")], axis=1)
        out_base = meta.join(agg, how="left")

        # Review DF (KEEP only) — mark only the cutoff row
        df = review_df.copy()
        df["YearMonth"] = pd.to_datetime(df["YearMonth"]).dt.to_period("M").dt.to_timestamp(how="start")
        is_keep = (df.get("Group", pd.Series(index=df.index, dtype=str)) == "000")
        if "source" in df.columns:
            is_keep = is_keep | (df["source"].astype(str).str.upper() == "KEEP")
        df_cut = df["YearMonth"].eq(cutoff)
        
        if target_scope.upper() == "ALL":
            df_target = df.loc[df_cut, ["MATERIAL_KEY"]].drop_duplicates().set_index("MATERIAL_KEY")
        else:
            # Existing behavior (KEEP only)
            is_keep = (df.get("Group", pd.Series(index=df.index, dtype=str)) == "000")
            if "source" in df.columns:
                is_keep = is_keep | (df["source"].astype(str).str.upper() == "KEEP")
            df_target = df.loc[df_cut & is_keep.fillna(False), ["MATERIAL_KEY"]].drop_duplicates().set_index("MATERIAL_KEY")

        out = df_target.join(out_base, how="left")

        # Default columns / initial values
        out["FurtherReviewNeeded"] = 0
        out["FR_TotalDiffRatio"] = np.nan
        out["FR_Reason"] = "ok"

        # ── Always define the 'not-evaluable/blocked' masks ──
        mask_no_cov = out["Covered_Months"].isna() | (out["Covered_Months"] == 0)

        # Set defaults so the masks always exist
        mask_incomplete = pd.Series(False, index=out.index)

        if require_full:
            # Require 100% completeness
            mask_incomplete = ~out["SF_Complete"].fillna(False)
            out.loc[mask_incomplete & ~mask_no_cov, "FR_Reason"] = "incomplete_sf"
        else:
            # Partial allowed: treat insufficient coverage ratio as 'incomplete'
            mask_incomplete = out["Coverage_Ratio"].fillna(0) < float(min_coverage_ratio)
            out.loc[mask_incomplete & ~mask_no_cov, "FR_Reason"] = "low_coverage"

        # Evaluable
        mask_eval = (~mask_no_cov) & (~mask_incomplete)
        mask_eval &= out["Model_Sum"].notna() & out["SF_Sum"].notna()

        denom = out.loc[mask_eval, "SF_Sum"].replace(0, np.nan)
        ratio = (out.loc[mask_eval, "Model_Sum"] - out.loc[mask_eval, "SF_Sum"]).abs() / denom

        out.loc[mask_eval, "FR_TotalDiffRatio"] = ratio
        out.loc[mask_eval & (ratio > diff_ratio_thr), "FurtherReviewNeeded"] = 1

        # Additional reasons
        out.loc[out["Model_Sum"].isna() & ~mask_incomplete & ~mask_no_cov, "FR_Reason"] = "no_model"
        out.loc[out["SF_Sum"].isna()    & ~mask_incomplete & ~mask_no_cov, "FR_Reason"] = "no_sf"


        # Write back into the review DF (cutoff row only)
        review_df = review_df.copy()
        review_df["FurtherReviewNeeded"] = 0
        review_df["FR_TotalDiffRatio"] = np.nan
        review_df["FR_Reason"] = "n/a"

        merged = review_df.loc[df_cut, ["MATERIAL_KEY"]].join(out, how="left", on="MATERIAL_KEY")
        review_df.loc[df_cut, "FurtherReviewNeeded"] = merged["FurtherReviewNeeded"].fillna(0).astype(int).values
        review_df.loc[df_cut, "FR_TotalDiffRatio"] = merged["FR_TotalDiffRatio"].values
        review_df.loc[df_cut, "FR_Reason"] = merged["FR_Reason"].fillna("n/a").values
        return review_df
    # ===== /Further Review Needed helper =====


    # ---------------- E2E REVIEW ----------------
    def e2e_review(progress_cb=None):
        upd = progress_cb or (lambda p, m: None)
        upd(2, "Initialize…")

        df_fcst_all = sf_fetch_all(sf_user, sf_acct, sf_wh, sf_db, sf_schema, sf_auth, brand)
        upd(8, "Snowflake data cached")

        # Bind Snowflake loader function for M1
        def _m1_load_fcst_from_snowflake(brand: str = "WG"):
            return df_fcst_all.copy()
        M1.load_fcst_from_snowflake = _m1_load_fcst_from_snowflake

        # A) Grouping (M1) — pass sidebar flag thresholds
        upd(12, "Calculating Groups …")
        review_df = M1.run_pipeline(
            pos_path=pos_path, brand=brand, cutoff=cutoff_ts,
            flag_thr1=flag_thr1, flag_thr2=flag_thr2, flag_thr3=flag_thr3,
            compare_months=compare_months
        )

        review_df["MATERIAL_KEY"] = canon_key_series(review_df["MATERIAL_KEY"])
        review_groups = review_df[["MATERIAL_KEY","Group"]].drop_duplicates()
        upd(18, "Complete")

        # ---- Shared precomputed (run once) ----
        upd(22, "POS preprocessing…")
        pos_raw_tmp = M1.read_pos_raw(pos_path, header_row=M1.RAW_POS_HEADER_ROW)
        monthly_tmp = M2.convert_weekly_to_monthly_long(
            M2.pos_preprocess(pos_raw_tmp, brand=brand),
            brand=brand, cutoff_for_missing_year=cutoff_ts
        )
        if "Segment" in monthly_tmp.columns:
            monthly_tmp["Segment"] = normalize_segment(monthly_tmp["Segment"]).replace(SEGMENT_CANON_MAP)
        monthly_tmp = M2.add_segment_dummies(monthly_tmp)
        monthly_tmp["Material"] = canon_key_series(monthly_tmp["Material"])
        df_cut = monthly_tmp.loc[monthly_tmp["YearMonth"] <= cutoff_ts].copy()

        # 🔹 Monthly Sellout = UPM × Door × WEEKS_PM (month before cutoff)
        sell_cols = ["Material","YearMonth","UPM","Door"]
        have = [c for c in sell_cols if c in df_cut.columns]
        if set(["Material","YearMonth","UPM","Door"]).issubset(have):
            pos_sellout = (df_cut[sell_cols]
                           .assign(SELLOUT=lambda x: x["UPM"] * x["Door"] * float(M1.WEEKS_PM)))
            pos_sellout = (pos_sellout[["Material","YearMonth","SELLOUT"]]
                           .rename(columns={"Material":"MATERIAL_KEY","YearMonth":"ds"}))
            pos_sellout["MATERIAL_KEY"] = canon_key_series(pos_sellout["MATERIAL_KEY"])
            pos_sellout["ds"] = pos_sellout["ds"].apply(period_start)
        else:
            pos_sellout = pd.DataFrame(columns=["MATERIAL_KEY","ds","SELLOUT"])

        upd(28, "Long/Short SKU Detection + Door Count Calculation…")
        span = (df_cut.dropna(subset=["UPM"]).groupby("Material", as_index=False)["YearMonth"]
                .nunique().rename(columns={"YearMonth":"n_months"}))

        door_exact = (monthly_tmp.loc[monthly_tmp["YearMonth"] == cutoff_ts, ["Material","Door"]]
                      .groupby("Material", as_index=False)
                      .max()
                      .rename(columns={"Door": "door_at_cutoff"}))

        sel_tmp = span.merge(door_exact, on="Material", how="left")
        sel_tmp["door_at_cutoff"] = pd.to_numeric(sel_tmp["door_at_cutoff"], errors="coerce").fillna(0.0)

        long_list  = sel_tmp.loc[(sel_tmp["n_months"] >= M2.MIN_HISTORY) & (sel_tmp["door_at_cutoff"] > 0), "Material"].tolist()
        short_list = sel_tmp.loc[(sel_tmp["n_months"] <  M2.MIN_HISTORY) & (sel_tmp["door_at_cutoff"] > 0), "Material"].tolist()




        # df_fcst_all: full data already fetched from Snowflake (MATERIAL_KEY, YearMonth, ESHIP, ASHIP)
        # df_cut: POS→monthly preprocessed result (Material, YearMonth, UPM, Door, ...)

        # 1) Align keys
        df_fcst_all2 = df_fcst_all.copy()
        df_fcst_all2["MATERIAL_KEY"] = canon_key_series(df_fcst_all2["MATERIAL_KEY"])
        df_fcst_all2["YearMonth"]    = normalize_yearmonth(df_fcst_all2["YearMonth"])

        # 2) Align to Material key (App df_cut uses Material; Snowflake uses MATERIAL_KEY)
        key_map = df_cut[["Material"]].drop_duplicates().copy()
        key_map["MATERIAL_KEY"] = key_map["Material"]

        # 3) Attach only needed columns (if ASHIP-only, just this)
        aship_ts = (
            df_fcst_all2[["MATERIAL_KEY","YearMonth","ASHIP"]]
            .merge(key_map, on="MATERIAL_KEY", how="inner")
            .drop(columns=["MATERIAL_KEY"])
        )

        # 4) Monthly merge into df_cut
        df_cut = df_cut.merge(aship_ts, on=["Material","YearMonth"], how="left")

        upd(35, "Calculating Seasonality…")
        seas_long = M3.build_long_seasonality(df_cut, long_list)
        seas_seg  = M3.build_segment_seasonality(df_cut)
        precomputed = {"df_cut": df_cut, "long_list": long_list, "short_list": short_list,
                       "seas_long": seas_long, "seas_seg": seas_seg}
        upd(42, "Preprocessing Complete")

        # B) USW FCST(M2)
        upd(50, "FCSTing (U/S/W)…")
        usw_fcst, usw_info = M2.run_usw_fcst(
            pos_path=pos_path, usw_path=usw_path, brand=brand,
            cutoff=cutoff_ts, periods=int(periods),
            drop_missing_baseline=False, drop_zero_baseline=False,
            precomputed=precomputed,
        )
        usw_fcst["MATERIAL_KEY"] = canon_key_series(usw_fcst["Material"])
        usw_fcst["ds"] = usw_fcst["ds"].apply(period_start)
        usw_fcst = usw_fcst.rename(columns={
            "yhat_denorm":   "USW_FCST",
            "factor":        "USW_factor",
            "baseline_units":"USW_baseline",
            "USW":           "USW_raw_per_week",
            "USW_month":     "USW_per_month",
        })
        usw_base_lookup = (usw_fcst[["Material","USW_baseline"]].drop_duplicates("Material").copy())
        usw_base_lookup["Material"] = canon_key_series(usw_base_lookup["Material"])
        usw_keep = safe_take(usw_fcst, [
            "MATERIAL_KEY","ds","Segment","method",
            "USW_factor","USW_baseline","USW_FCST",
            "USW_raw_per_week","USW_per_month","Door_at_cutoff"
        ]).rename(columns={"Segment":"Segment_usw","method":"method_usw"})
        upd(62, "FCST U/S/W Complete")

        # C) ASHIP FCST(M3)
        upd(68, "FCSTing (ASHIP & U/S/W)…")
        aship_fcst, aship_info = M3.run_aship_fcst(
            pos_path=pos_path, brand=brand,
            cutoff=cutoff_ts, periods=int(periods),
            df_fcst=df_fcst_all[["MATERIAL_KEY","YearMonth","ASHIP"]].copy(),
            drop_missing_baseline=False,
            drop_zero_baseline=False, strict_three_months=True,
            precomputed=precomputed,
            usw_baseline_lookup=usw_base_lookup,
            usw_weight=0.50,
        )
        aship_fcst["MATERIAL_KEY"] = canon_key_series(aship_fcst["Material"])
        aship_fcst["ds"] = aship_fcst["ds"].apply(period_start)
        aship_fcst = aship_fcst.rename(columns={
            "yhat_denorm":"ASHIP_FCST",
            "factor":"ASHIP_factor",
            "baseline_units":"M3_baseline"
        })
        aship_keep = aship_fcst[[
            "MATERIAL_KEY","ds","Segment","method","ASHIP_factor","M3_baseline","ASHIP_FCST"
        ]].rename(columns={"Segment":"Segment_aship","method":"method_aship"})
        upd(78, "FCST Complete")

        # D~F) Merge & select
        upd(85, "Merging & Selection…")

        cal = (pd.concat([
                aship_keep[["MATERIAL_KEY","ds"]],
                usw_keep[["MATERIAL_KEY","ds"]] if not usw_keep.empty else aship_keep[["MATERIAL_KEY","ds"]],
            ], axis=0).drop_duplicates().sort_values(["MATERIAL_KEY","ds"])
        )
        cal = cal.merge(review_groups, on="MATERIAL_KEY", how="left")
        cal = (cal.merge(aship_keep, on=["MATERIAL_KEY","ds"], how="left")
                 .merge(usw_keep,   on=["MATERIAL_KEY","ds"], how="left"))

        if "Segment" not in cal.columns:
            cal["Segment"] = np.nan
        if "Segment_aship" in cal.columns:
            cal["Segment"] = cal["Segment"].fillna(cal["Segment_aship"])
        if "Segment_usw" in cal.columns:
            cal["Segment"] = cal["Segment"].fillna(cal["Segment_usw"])

        seg_map_all = (precomputed["df_cut"][["Material","Segment"]]
                       .dropna()
                       .drop_duplicates("Material"))
        seg_map_all["MATERIAL_KEY"] = canon_key_series(seg_map_all["Material"])
        seg_map_all = seg_map_all.rename(columns={"Segment": "Segment_map"})
        cal = cal.merge(seg_map_all[["MATERIAL_KEY","Segment_map"]],
                        on="MATERIAL_KEY", how="left")
        cal["Segment"] = cal["Segment"].fillna(cal["Segment_map"])
        cal = cal.drop(columns=["Segment_map"])

        cal["Segment"] = normalize_segment(cal["Segment"]).replace(SEGMENT_CANON_MAP)

        
        def choose(row):
            # 1) If ASHIP exists, always select it (ignore group)
            if pd.notna(row.get("ASHIP_FCST")):
                return (row["ASHIP_FCST"], "ASHIP",
                        row.get("ASHIP_factor"), row.get("M3_baseline"),
                        row.get("Segment"), row.get("method_aship"))

            # 2) If no ASHIP, fall back to USW
            if pd.notna(row.get("USW_FCST")):
                return (row["USW_FCST"], "USW",
                        row.get("USW_factor"), row.get("USW_baseline"),
                        row.get("Segment"), row.get("method_usw"))

            # 3) If neither, KEEP
            return (np.nan, "KEEP", np.nan, np.nan, row.get("Segment"), None)
        

        chosen = [choose(r) for _, r in cal.iterrows()]
        cal["FCST"]              = [c[0] for c in chosen]
        cal["source"]            = [c[1] for c in chosen]
        cal["SeasonalityFactor"] = [c[2] for c in chosen]
        cal["BaselineUnits"]     = [c[3] for c in chosen]
        cal["Segment"]           = [c[4] for c in chosen]
        cal["LongShort"]         = [c[5] for c in chosen]

        cal["UPM_FCST"] = cal["USW_FCST"] if "USW_FCST" in cal.columns else np.nan
        cal["ASHIP_UPM_BLEND_FCST"] = cal["ASHIP_FCST"] if "ASHIP_FCST" in cal.columns else np.nan

        cal["Selected_FCST_before_trend"] = cal["FCST"]
        cal["Month"] = cal["ds"].dt.to_period("M").astype(str)

        final_fcst = (cal[[
            "MATERIAL_KEY","Group","Segment","LongShort",
            "ds","Month","source",
            "SeasonalityFactor","BaselineUnits",
            "UPM_FCST","ASHIP_UPM_BLEND_FCST",
            "ASHIP_FCST","ASHIP_factor","M3_baseline",
            "USW_FCST","USW_factor","USW_baseline",
            "USW_raw_per_week","USW_per_month","Door_at_cutoff",
            "Selected_FCST_before_trend"
        ]].sort_values(["MATERIAL_KEY","ds"]).reset_index(drop=True))
        upd(92, "Selection Complete")

        # G) Final-stage trend application
        final_fcst["ds"] = pd.to_datetime(final_fcst["ds"]).dt.to_period("M").dt.to_timestamp(how="start")

        if use_trend:
            upd(95, "Applying Trend…")
            materials = final_fcst["MATERIAL_KEY"].dropna().unique().tolist()
            trend_tbl = M3.compute_trend_table(
                df_cut=precomputed["df_cut"],
                materials=materials,
                cutoff=cutoff_ts,
                periods=int(periods),
                lookback=TREND_LOOKBACK_FIXED,
                ramp_months=TREND_RAMP_FIXED,
                up_pct=TREND_UP_FIXED,
                down_pct=TREND_DOWN_FIXED,
            ).rename(columns={"Material":"MATERIAL_KEY"})
            trend_tbl["ds"] = pd.to_datetime(trend_tbl["ds"]).dt.to_period("M").dt.to_timestamp(how="start")

            final_fcst = final_fcst.merge(trend_tbl, on=["MATERIAL_KEY","ds"], how="left")
            if "trend_multiplier" not in final_fcst.columns:
                final_fcst["trend_multiplier"] = 1.0
            else:
                final_fcst["trend_multiplier"] = final_fcst["trend_multiplier"].fillna(1.0)
        else:
            final_fcst["trend_multiplier"] = 1.0

        # --- KEEP(Trend) patch
        mask_keep   = final_fcst["source"].eq("KEEP")
        mask_trend  = final_fcst["trend_multiplier"].astype(float).ne(1.0)
        mask_usw_ok = final_fcst["USW_factor"].notna() & final_fcst["USW_baseline"].notna()
        use_idx = mask_keep & mask_trend & mask_usw_ok

        final_fcst.loc[use_idx, "Selected_FCST_before_trend"] = (
            final_fcst.loc[use_idx, "USW_factor"].astype(float)
            * final_fcst.loc[use_idx, "USW_baseline"].astype(float)
        )
        final_fcst.loc[use_idx, "SeasonalityFactor"] = final_fcst.loc[use_idx, "USW_factor"]
        final_fcst.loc[use_idx, "BaselineUnits"]     = final_fcst.loc[use_idx, "USW_baseline"]
        final_fcst.loc[use_idx, "source"] = "KEEP (Trend)"

        final_fcst["FCST"]       = final_fcst["Selected_FCST_before_trend"].astype(float)
        final_fcst["FCST_final"] = final_fcst["FCST"] * final_fcst["trend_multiplier"].astype(float)




        # === Backorder Buffer (trigger when ASHIP_4M > ESHIP_4M) ==============
        # === 4M averages & Buffer Rule (ASHIP vs SELLOUT) ==========================
        # (1) Define 4-month window (latest 4 months including cutoff)
        _4m_months = pd.period_range(
            (cutoff_ts - pd.DateOffset(months=3)).to_period("M"),
            cutoff_ts.to_period("M"),
            freq="M"
        ).to_timestamp(how="start")

        # (2) ESHIP/ASHIP 4M average (reuse Snowflake aggregation)
        _4m = df_fcst_all[df_fcst_all["YearMonth"].isin(_4m_months)].copy()
        _4m["MATERIAL_KEY"] = canon_key_series(_4m["MATERIAL_KEY"])
        _4m["YearMonth"]    = normalize_yearmonth(_4m["YearMonth"])
        _agg = (
            _4m.groupby("MATERIAL_KEY", as_index=False)
            .agg(ESHIP_4M=("ESHIP","mean"),
                    ASHIP_4M=("ASHIP","mean"))
        )

        # (3) SELLOUT 4M average (use pos_sellout built from POS→monthly)
        sell4m = pd.DataFrame(columns=["MATERIAL_KEY","SELLOUT_4M"])
        if pos_sellout is not None and not pos_sellout.empty:
            _sell = pos_sellout.copy()
            _sell["ds"] = pd.to_datetime(_sell["ds"], errors="coerce")
            _sell = _sell[_sell["ds"].isin(_4m_months)]
            if not _sell.empty and {"MATERIAL_KEY","SELLOUT"}.issubset(_sell.columns):
                sell4m = (
                    _sell.groupby("MATERIAL_KEY", as_index=False)
                        .agg(SELLOUT_4M=("SELLOUT","mean"))
                )

        # (4) Merge ESHIP/ASHIP 4M and SELLOUT 4M into final_fcst
        final_fcst = final_fcst.merge(_agg,   on="MATERIAL_KEY", how="left")
        final_fcst = final_fcst.merge(sell4m, on="MATERIAL_KEY", how="left")

        # (5) Apply buffer rule
        #     - If ASHIP_4M > SELLOUT_4M → add buffer (×1.10)
        #     - Else (SELLOUT_4M >= ASHIP_4M) → no buffer (×1.00)
        BUFFER_PCT = 0.15

        aship4m = pd.to_numeric(final_fcst["ASHIP_4M"], errors="coerce")
        sell4m_ = pd.to_numeric(final_fcst["SELLOUT_4M"], errors="coerce")

        need_buf = (aship4m.notna() & sell4m_.notna() & (aship4m > sell4m_))
        final_fcst["buffer_rule_multiplier"] = np.where(need_buf, 1.0 + BUFFER_PCT, 1.0)

        # Preserve the initially computed FCST_final, then apply the final buffered value
        if "FCST_final_base" not in final_fcst.columns:
            final_fcst["FCST_final_base"] = final_fcst["FCST_final"]

        final_fcst["FCST_final"] = final_fcst["FCST_final"] * final_fcst["buffer_rule_multiplier"]

        # (Optional) If you want to display whether buffer was applied:
        final_fcst["BufferApplied"] = final_fcst["buffer_rule_multiplier"].map(lambda x: bool(x and x > 1.0))


        # ── Attach previous forecast (prefer ESHIP, else ASHIP)
        df_prev = df_fcst_all.copy()
        df_prev["YearMonth"]    = normalize_yearmonth(df_prev["YearMonth"])
        df_prev["MATERIAL_KEY"] = canon_key_series(df_prev["MATERIAL_KEY"])
        df_prev = df_prev.rename(columns={"YearMonth": "ds"})
        h_start = cutoff_ts + pd.offsets.MonthBegin(1)
        h_end   = cutoff_ts + pd.DateOffset(months=int(periods))
        df_prev = df_prev[(df_prev["ds"] >= h_start) & (df_prev["ds"] <= h_end)]

        df_prev["previous_FCST"] = df_prev["ESHIP"]
        df_prev.loc[df_prev["previous_FCST"].isna(), "previous_FCST"] = df_prev["ASHIP"]
        df_prev["prev_source"] = np.where(
            df_prev["ESHIP"].notna(), "ESHIP",
            np.where(df_prev["ASHIP"].notna(), "ASHIP", "NONE")
        )

        prev_cols = ["MATERIAL_KEY", "ds", "previous_FCST", "prev_source"]
        final_fcst = final_fcst.merge(df_prev[prev_cols], on=["MATERIAL_KEY","ds"], how="left")

        final_fcst["diff_abs"] = final_fcst["FCST_final"] - final_fcst["previous_FCST"]
        final_fcst["diff_pct"] = np.where(
            final_fcst["previous_FCST"] > 0,
            (final_fcst["FCST_final"] / final_fcst["previous_FCST"]) - 1.0,
            np.nan
        )

        ################## Gross Change (first up-to 4M; require ≥2M; KEEP untouched)
        GROSS_CHANGE_THR = float(gross_thr)

        for _col, _default in [
            ("HasPrev", False),
            ("GrossChangePct", np.nan),
            ("GrossChangePct_view", "—"),
            ("FCST_Update_Needed", False),
        ]:
            if _col not in final_fcst.columns:
                final_fcst[_col] = _default

        H = int(min(4, int(periods)))

        if H >= 1:
            w_start = cutoff_ts + pd.offsets.MonthBegin(1)
            w_end   = cutoff_ts + pd.DateOffset(months=H)

            update_groups = {"001","011","101","111"}
            mask_update_target = (
                final_fcst["Group"].astype(str).isin(update_groups)
                | final_fcst["source"].eq("KEEP (Trend)")
            )

            scope = final_fcst.loc[
                (final_fcst["ds"] >= w_start) & (final_fcst["ds"] <= w_end) & mask_update_target,
                ["MATERIAL_KEY","FCST_final","previous_FCST","ds"]
            ].copy()

            if not scope.empty:
                agg = (
                    scope.groupby("MATERIAL_KEY", as_index=False)
                        .agg(
                            FCST_sum=("FCST_final",  lambda s: pd.to_numeric(s, errors="coerce").sum(min_count=1)),
                            Prev_sum=("previous_FCST", lambda s: pd.to_numeric(s, errors="coerce").sum(min_count=1)),
                            n_months=("ds", "nunique"),
                        )
                )

                agg["HasPrev"] = (agg["n_months"] >= 2) & (pd.to_numeric(agg["Prev_sum"], errors="coerce") > 0)

                def _chg(row):
                    if not row["HasPrev"]:
                        return np.nan
                    return (float(row["FCST_sum"]) / float(row["Prev_sum"])) - 1.0

                agg["GrossChangePct"] = agg.apply(_chg, axis=1)

                final_fcst = final_fcst.merge(
                    agg[["MATERIAL_KEY","GrossChangePct","HasPrev"]],
                    on="MATERIAL_KEY", how="left", suffixes=("", "_calc")
                )
                for c in ["HasPrev","GrossChangePct"]:
                    calc = f"{c}_calc"
                    if calc in final_fcst.columns:
                        final_fcst[c] = final_fcst[calc].where(final_fcst[calc].notna(), final_fcst[c])
                        final_fcst.drop(columns=[calc], inplace=True, errors="ignore")

        final_fcst["HasPrev"] = final_fcst["HasPrev"].fillna(False)
        final_fcst["GrossChangePct_view"] = np.where(
            final_fcst["HasPrev"] & np.isfinite(final_fcst["GrossChangePct"]),
            (final_fcst["GrossChangePct"] * 100).round(1).astype(str) + "%", 
            "—"
        )
        final_fcst["FCST_Update_Needed"] = (
            final_fcst["HasPrev"]
            & np.isfinite(final_fcst["GrossChangePct"])
            & (final_fcst["GrossChangePct"].abs() >= GROSS_CHANGE_THR)
        )
        ################## end Gross Change ####################

        upd(100, "Final FCST Calculation Complete")
        return review_df, usw_fcst, aship_fcst, final_fcst, usw_info, aship_info, df_fcst_all, pos_sellout

    # ---------------- Execute / save to session + render ----------------
    try:
        if run_btn:
            with st.spinner("Running…"):
                review_df, usw_fcst, aship_fcst, final_fcst, usw_info, aship_info, df_fcst_all, pos_sellout = e2e_review(progress_cb=progress_cb)
                
            progress_done("FCST Review complete!")
            st.success("Success!")

            sf_fcst = None
            try:
                if ('df_fcst_all' in locals()) and (df_fcst_all is not None) and (not df_fcst_all.empty):
                    # Horizon range (horizon_months including cutoff)
                    months = pd.period_range(
                        cutoff_ts.to_period("M"),
                        (cutoff_ts.to_period("M") + (int(fr_months) - 1)),
                        freq="M"
                    ).to_timestamp(how="start")

                    tmp = df_fcst_all.copy()

                    # Normalize YearMonth to month-start datetime
                    tmp["YearMonth"] = pd.to_datetime(tmp["YearMonth"], errors="coerce")
                    # If YearMonth is an integer YYYYMM, the line above may produce NaT; handle that case:
                    if tmp["YearMonth"].isna().any():
                        # Try safe parsing for YYYYMM format
                        ym_raw = df_fcst_all["YearMonth"].astype(str).str.replace(r"[^0-9]", "", regex=True)
                        tmp["YearMonth"] = pd.to_datetime(ym_raw + "01", format="%Y%m%d", errors="coerce")
                    # Unify to month start
                    tmp["YearMonth"] = tmp["YearMonth"].dt.to_period("M").dt.to_timestamp(how="start")

                    # Normalize MATERIAL_KEY (string, trim)
                    if "MATERIAL_KEY" in tmp.columns:
                        tmp["MATERIAL_KEY"] = tmp["MATERIAL_KEY"].astype(str).str.strip()
                    else:
                        raise KeyError("df_fcst_all에 MATERIAL_KEY 컬럼이 없습니다.")

                    # Slice to target months only
                    tmp = tmp[tmp["YearMonth"].isin(months)]

                    # Snowflake_FCST = prefer ESHIP, else ASHIP
                    if "ESHIP" not in tmp.columns:
                        raise KeyError("df_fcst_all에 ESHIP 컬럼이 없습니다.")
                    if "ASHIP" not in tmp.columns:
                        raise KeyError("df_fcst_all에 ASHIP 컬럼이 없습니다.")

                    tmp["Snowflake_FCST"] = pd.to_numeric(tmp["ESHIP"], errors="coerce")
                    need_fill = tmp["Snowflake_FCST"].isna()
                    tmp.loc[need_fill, "Snowflake_FCST"] = pd.to_numeric(tmp.loc[need_fill, "ASHIP"], errors="coerce")

                    sf_fcst = tmp[["MATERIAL_KEY","YearMonth","Snowflake_FCST"]].copy()

                    # If completely empty, set to None (so the next step auto-skips)
                    if sf_fcst.empty:
                        sf_fcst = None
            except Exception as e:
                # If building fails, quietly skip (optionally add st.warning(str(e)))
                sf_fcst = None


            if 'sf_fcst' in locals() and sf_fcst is not None and not sf_fcst.empty:
                review_df = compute_further_review_flags(
                    review_df=review_df,
                    final_fcst=final_fcst,
                    sf_fcst=sf_fcst,
                    cutoff=cutoff_ts,
                    horizon_months=int(fr_months),   # e.g., 4
                    diff_ratio_thr=float(fr_thr),
                    model_col="FCST_final",
                    sf_col="Snowflake_FCST",
                    require_full=False,              # ← Allow partial coverage
                    min_coverage_ratio=0.75,         # ← ≥3 out of 4 months (0.75)
                )

            else:
                # If no Snowflake forecast, skip computation (keep review table as-is)
                pass

            try:
                _cut = cutoff_ts.to_period("M").to_timestamp(how="start")
                _fr_src = review_df.copy()
                _fr_src["YearMonth"] = pd.to_datetime(_fr_src["YearMonth"], errors="coerce")\
                                        .dt.to_period("M").dt.to_timestamp(how="start")
                fr_items = (_fr_src.query("YearMonth == @_cut and FurtherReviewNeeded == 1")
                                ["MATERIAL_KEY"].dropna().astype(str).unique().tolist())
                final_fcst["FR_Needed"] = final_fcst["MATERIAL_KEY"].astype(str).isin(fr_items)
            except Exception:
                # Safety net: even if this fails, keep the app running
                final_fcst["FR_Needed"] = False

            st.session_state.review_results = {
                "review_df": review_df,
                "usw_fcst": usw_fcst,
                "aship_fcst": aship_fcst,
                "final_fcst": final_fcst,
                "usw_info": usw_info,
                "aship_info": aship_info,
                "df_fcst_all": df_fcst_all,
                "pos_sellout": pos_sellout,
                "brand": brand,
                "cutoff_ts": cutoff_ts,
            }
            st.session_state.ran_review = True

        # Load from session
        R = st.session_state.review_results

        # ★★★ Apply manual % adjustments
        def apply_manual_adjustments(df: pd.DataFrame, adj_dict: dict | None = None) -> pd.DataFrame:
            df = df.copy()
            if "FCST_final_base" not in df.columns:
                df["FCST_final_base"] = df["FCST_final"]

            use_dict = adj_dict if adj_dict is not None else st.session_state.fcst_adj

            keys = []
            for _, r in df.iterrows():
                k = (str(r["MATERIAL_KEY"]),
                    pd.to_datetime(r["ds"]).to_period("M").to_timestamp(how="start"))
                keys.append(k)

            pct = np.array([float(use_dict.get(k, 0.0)) for k in keys], dtype=float)
            mul = 1.0 + pct

            df["manual_adj_multiplier"] = mul
            df["manual_adj_pct"] = pct
            df["FCST_final"] = df["FCST_final_base"] * df["manual_adj_multiplier"]

            if "previous_FCST" in df.columns:
                prev = pd.to_numeric(df["previous_FCST"], errors="coerce")
                df["diff_abs"] = df["FCST_final"] - prev
                df["diff_pct"] = np.where(prev > 0, (df["FCST_final"] / prev) - 1.0, np.nan)
            return df

        # ★★★ Apply Update Needed overrides
        def apply_update_needed_overrides(df: pd.DataFrame) -> pd.DataFrame:
            df = df.copy()
            base = df.get("FCST_Update_Needed", pd.Series(False, index=df.index)).fillna(False)
            eff = base.copy()

            ov = st.session_state.update_needed_overrides
            if ov:
                for i, r in df.iterrows():
                    try:
                        mat = str(r.get("MATERIAL_KEY", r.get("Material", "")))
                        ds_key = pd.to_datetime(r["ds"]).to_period("M").to_timestamp(how="start")
                        k = (mat, ds_key)
                        if k in ov:
                            eff.iat[i] = bool(ov[k])
                    except Exception:
                        pass
            df["Update_Needed_eff"] = eff
            return df

        def render_global_graph(R, cutoff_ts: pd.Timestamp):
            st.divider()
            st.markdown("### 📊 Visualization ; SKU history & FCST")

            final_fcst  = R["final_fcst"]
            df_fcst_all = R["df_fcst_all"]
            pos_sellout = R.get("pos_sellout", pd.DataFrame())

            # ─────────────────────────────────────────────────────────
            # Graph target candidates (prioritize items needing update)
            # ─────────────────────────────────────────────────────────
            agg_for_graph = (
                final_fcst.groupby("MATERIAL_KEY", as_index=False)
                        .agg(
                            any_update=("FCST_Update_Needed", lambda s: bool(pd.Series(s).fillna(False).any())),
                            max_abs_gross=("GrossChangePct", lambda s: float(pd.to_numeric(s, errors="coerce").abs().max()))
                        )
            )
            agg_for_graph["max_abs_gross"] = agg_for_graph["max_abs_gross"].fillna(0.0)
            agg_for_graph = agg_for_graph.sort_values(["any_update", "max_abs_gross"], ascending=[False, False])

            only_updates_for_picker = st.checkbox(
                "Show only items needing update (for graph picker)",
                value=True,
                key="graph_picker_only_updates"
            )
            mats_for_picker = (
                agg_for_graph.loc[agg_for_graph["any_update"], "MATERIAL_KEY"].tolist()
                if only_updates_for_picker else agg_for_graph["MATERIAL_KEY"].tolist()
            )
            selected_mats = st.multiselect(
                "Select Items 純愛",
                mats_for_picker,
                default=mats_for_picker,
                key="graph_picker_selected_mats"
            )

            # ─────────────────────────────────────────────────────────
            # Internal utilities
            # ─────────────────────────────────────────────────────────
            def _build_hist_df(mat_key: str) -> pd.DataFrame:
                hist = df_fcst_all[df_fcst_all.get("MATERIAL_KEY", pd.Series(dtype=str)) == mat_key].copy()
                if "YearMonth" in hist.columns:
                    hist = hist.rename(columns={"YearMonth": "ds"})
                hist["ds"] = pd.to_datetime(hist["ds"], errors="coerce")
                hist = hist[hist["ds"] <= cutoff_ts]

                have_cols = [c for c in ["ESHIP", "ASHIP"] if c in hist.columns]
                if have_cols:
                    hist = hist.groupby("ds", as_index=False)[have_cols].sum(min_count=1)
                else:
                    hist = hist[["ds"]].drop_duplicates()

                # POS Sellout (if available)
                if pos_sellout is not None and not pos_sellout.empty:
                    sell = pos_sellout[pos_sellout.get("MATERIAL_KEY", pd.Series(dtype=str)) == mat_key]
                    if {"ds", "SELLOUT"}.issubset(sell.columns):
                        sell = sell[["ds", "SELLOUT"]].copy()
                        sell["ds"] = pd.to_datetime(sell["ds"], errors="coerce")
                        hist = hist.merge(sell, on="ds", how="outer").sort_values("ds")
                return hist

            def _build_future_pairs(mat_key: str, preview_ratio: float):
                """
                반환:
                - fut_base    : 프리뷰 미반영(=현재 확정된 FCST_final vs previous)
                - fut_preview : 프리뷰 반영(=전달된 preview_ratio를 곱한 가상 FCST_final)
                - fut_ds_list : 컷오프 이후 월 리스트
                """
                fut_all = R["final_fcst"].copy()
                fut_all["ds"] = pd.to_datetime(fut_all["ds"], errors="coerce")
                h_start = cutoff_ts + pd.offsets.MonthBegin(1)

                # Data for confirmed (solid line)
                base = fut_all[
                    (fut_all["MATERIAL_KEY"] == mat_key) & (fut_all["ds"] >= h_start)
                ][["ds", "previous_FCST", "FCST_final"]].copy()

                fut_ds_list = base["ds"].dt.to_period("M").dt.to_timestamp(how="start").tolist()

                # For preview (dashed): multiply base FCST_final by preview_ratio to change display only
                prevv = base.copy()
                prevv["FCST_final"] = np.rint(
                    pd.to_numeric(prevv["FCST_final"], errors="coerce") * float(preview_ratio)
                ).astype("Int64")

                return base, prevv, fut_ds_list

            # If nothing selected, exit
            if not selected_mats:
                st.info("No Item has been selected.")
                return

            # Tab header horizontal scroll CSS
            st.markdown("""
            <style>
            .stTabs [role="tablist"]{
            overflow-x: auto !important;
            overflow-y: hidden;
            white-space: nowrap;
            scrollbar-width: thin;
            }
            .stTabs [role="tab"]{
            flex: 0 0 auto !important;
            }
            .stTabs [role="tablist"]::-webkit-scrollbar { height: 6px; }
            .stTabs [role="tablist"]::-webkit-scrollbar-thumb { border-radius: 4px; }
            </style>
            """, unsafe_allow_html=True)

            tabs = st.tabs([str(m) for m in selected_mats])

            # ─────────────────────────────────────────────────────────
            # Single-SKU renderer (graph + preview buffer UI)
            # ─────────────────────────────────────────────────────────
            def _render_one_mat(mat_key: str):
                # ── 0) Read & store the preview buffer (%) slider first ─────────────────────────
                st.markdown("##### ✍️ Preview Buffer")
                # Default: if you want, you can set the default value using the existing preview/confirmed adjustments; tweak the value logic below.
                # Here we use 0.0% as the default.
                glb = st.slider(
                    f"Global preview % — {mat_key}",
                    min_value=-50.0, max_value=100.0, value=0.0, step=1.0,
                    key=f"preview_global_{mat_key}"
                )
                preview_ratio = float(1.0 + glb / 100.0)
                st.session_state[f"preview_global_ratio_{mat_key}"] = preview_ratio

                # ── 1) Prepare left/right graph data (using the latest preview_ratio) ────────────
                hist = _build_hist_df(mat_key)
                fut_base, fut_prev, fut_ds_list = _build_future_pairs(mat_key, preview_ratio)

                # Left (history) melt
                left_df = pd.DataFrame()
                if not hist.empty:
                    value_vars = [c for c in ["SELLOUT", "ESHIP", "ASHIP"] if c in hist.columns]
                    if value_vars:
                        left_df = (
                            hist.melt(id_vars="ds", value_vars=value_vars, var_name="Type", value_name="Value")
                                .dropna(subset=["Value"])
                        )

                # Right (post-cutoff): confirmed (solid) + preview (dashed)
                right_df = pd.DataFrame()
                if not fut_base.empty or not fut_prev.empty:
                    # Confirmed (solid)
                    right_base = (
                        fut_base.rename(columns={"previous_FCST": "Previous_FCST", "FCST_final": "Proposed_FCST"})
                                .melt(id_vars="ds", value_vars=["Previous_FCST", "Proposed_FCST"],
                                    var_name="Type", value_name="Value")
                    )
                    # Preview (dashed)
                    right_prev = pd.DataFrame()
                    if not fut_prev.empty:
                        right_prev = (
                            fut_prev.rename(columns={"FCST_final": "Proposed_Preview"})
                                    .melt(id_vars="ds", value_vars=["Proposed_Preview"],
                                        var_name="Type", value_name="Value")
                        )

                    right_df = pd.concat([right_base, right_prev], ignore_index=True)
                    right_df = right_df.dropna(subset=["Value"])

                # Draw with combined left/right
                plot_df = pd.concat([left_df, right_df], ignore_index=True)
                if plot_df.empty:
                    st.info(f"{mat_key}: No data to display.")
                    return

                plot_df["ds"] = pd.to_datetime(plot_df["ds"])
                plot_df = plot_df.sort_values("ds")

                type_order = ["SELLOUT", "ESHIP", "ASHIP", "Previous_FCST", "Proposed_FCST", "Proposed_Preview"]
                fig = px.line(
                    plot_df, x="ds", y="Value", color="Type",
                    category_orders={"Type": type_order}, color_discrete_map=COLOR_MAP,
                    title=f"{mat_key} — History & FCST (Previous / Proposed / Preview)"
                )

                # Preview dashed-line style
                fig.update_traces(mode="lines+markers", marker=dict(size=8))
                for i, d in enumerate(fig.data):
                    if d.name == "Proposed_Preview":
                        fig.data[i].line.update(dash="dash")

                fig.update_layout(height=800, hovermode="x unified", hoverdistance=8)
                fig.update_traces(hovertemplate="<b>%{fullData.name}</b>: %{y:,.0f}<extra></extra>")
                fig.update_xaxes(spikesnap="data", hoverformat="%m/%Y")
                fig.update_yaxes(spikesnap="data")
                fig.add_vline(x=cutoff_ts, line_dash="dash", line_color="red")

                xmin = plot_df["ds"].min().to_period("M").to_timestamp(how="start")
                xmax = plot_df["ds"].max().to_period("M").to_timestamp(how="start")
                fig.update_xaxes(range=[xmin, xmax], dtick="M1", tickformat="%m/%Y", tick0=xmin, showgrid=True, title_text="Month")
                fig.update_yaxes(showgrid=True, title_text="Units")
                st.plotly_chart(fig, use_container_width=True)

                # ── History table ───────────────────────────────────────────────────────────────
                st.markdown("**History (Before Cutoff)**")
                if not hist.empty:
                    hist_tbl = hist.sort_values("ds").copy()
                    hist_tbl["Month"] = (
                        pd.to_datetime(hist_tbl["ds"], errors="coerce")
                        .dt.to_period("M")
                        .astype(str)
                    )
                    metrics = [c for c in ["SELLOUT","ESHIP","ASHIP"] if c in hist_tbl.columns]
                    show_cols = ["Month"] + metrics
                    hist_tbl = hist_tbl[show_cols].copy()

                    for c in metrics:
                        hist_tbl[c] = np.rint(pd.to_numeric(hist_tbl[c], errors="coerce")).astype("Int64")

                    hist_pivot = (
                        hist_tbl.set_index("Month")[metrics]
                                .T
                                .reset_index()
                                .rename(columns={"index":"Metric"})
                    )
                    st.dataframe(hist_pivot, use_container_width=True, height=220)
                else:
                    st.info("No historical data to display for this SKU.")

                # ── Post-cutoff summary ─────────────────────────────────────────────────────────
                st.markdown("**Summary (After Cutoff)**")
                if not fut_base.empty:
                    _sum = fut_base.sort_values("ds").copy()
                    _sum["Month"] = (
                        pd.to_datetime(_sum["ds"], errors="coerce")
                        .dt.to_period("M")
                        .astype(str)
                    )
                    metrics2 = ["previous_FCST", "FCST_final"]
                    for c in metrics2:
                        _sum[c] = np.rint(pd.to_numeric(_sum[c], errors="coerce")).astype("Int64")

                    sum_pivot = (
                        _sum[["Month"] + metrics2]
                            .set_index("Month")[metrics2]
                            .T
                            .reset_index()
                            .rename(columns={
                                "index": "Metric",
                                "previous_FCST": "FCST_previous",
                                "FCST_final": "FCST_final"
                            })
                    )
                    st.dataframe(sum_pivot, use_container_width=True, height=220)
                else:
                    st.info("No data after Cutoff.")

                # ── APPLY / RESET button area ───────────────────────────────────────────────────
                c_apply, c_reset_preview, c_reset_all = st.columns([1, 1, 1])

                # APPLY: convert the current preview_ratio into a 'confirmed adjustment'
                with c_apply:
                    if st.button(f"APPLY preview → final ({mat_key})", key=f"apply_tab_v1_{mat_key}"):
                        # Reuse the latest ratio from session (already saved above)
                        ratio_now = st.session_state.get(f"preview_global_ratio_{mat_key}", 1.0)

                        # Fetch the post-cutoff original values and baseline for the current mat_key
                        base_all = R["final_fcst"].copy()
                        base_all["ds"] = pd.to_datetime(base_all["ds"], errors="coerce").dt.to_period("M").dt.to_timestamp(how="start")

                        base_rows = base_all[
                            (base_all["MATERIAL_KEY"] == mat_key) & (base_all["ds"] >= cutoff_ts + pd.offsets.MonthBegin(1))
                        ]

                        # Virtual result: multiply the on-screen (pre-apply) values by the preview ratio
                        _apply_df = fut_base.copy()
                        _apply_df["ds"] = pd.to_datetime(_apply_df["ds"], errors="coerce").dt.to_period("M").dt.to_timestamp(how="start")

                        _apply_df["FCST_final_new"] = pd.to_numeric(_apply_df["FCST_final"], errors="coerce") * ratio_now

                        # Save the % change vs. baseline into session
                        for _, row in _apply_df.iterrows():
                            d = row["ds"]
                            base_row = base_rows.loc[base_rows["ds"] == d]
                            if not base_row.empty:
                                base_val = pd.to_numeric(
                                    base_row["FCST_final_base"] if "FCST_final_base" in base_row.columns else base_row["FCST_final"],
                                    errors="coerce"
                                ).squeeze()
                                new_val = pd.to_numeric(row["FCST_final_new"], errors="coerce")
                                if pd.notna(base_val) and base_val > 0 and pd.notna(new_val):
                                    pct = float(new_val / base_val - 1.0)   # Save (ratio - 1) in %
                                    st.session_state.fcst_adj[(mat_key, d)] = pct

                        # Clear preview traces and refresh
                        for d in fut_ds_list:
                            st.session_state.fcst_adj_preview.pop((mat_key, d), None)
                        st.rerun()

                with c_reset_preview:
                    if st.button(f"Reset Preview ({mat_key})", key=f"reset_preview_tab_{mat_key}"):
                        st.session_state[f"preview_global_ratio_{mat_key}"] = 1.0
                        st.rerun()

                with c_reset_all:
                    if st.button(f"Reset All ({mat_key})", key=f"reset_all_tab_{mat_key}"):
                        for d in fut_ds_list:
                            st.session_state.fcst_adj_preview.pop((mat_key, d), None)
                            st.session_state.fcst_adj.pop((mat_key, d), None)
                        st.rerun()


            # ─────────────────────────────────────────────────────────
            # Render tabs
            # ─────────────────────────────────────────────────────────
            for tab, mat in zip(tabs, selected_mats):
                with tab:
                    _render_one_mat(mat)
                


        if not R:
            st.info("왼쪽에서 파일/파라미터를 설정하고 RUN을 눌러주세요.")
            st.stop()

        review_df   = R["review_df"]
        usw_fcst    = R["usw_fcst"]
        aship_fcst  = R["aship_fcst"]
        final_fcst  = R["final_fcst"]
        usw_info    = R["usw_info"]
        aship_info  = R["aship_info"]
        df_fcst_all = R["df_fcst_all"]
        pos_sellout = R["pos_sellout"]
        brand       = R["brand"]
        cutoff_ts   = R["cutoff_ts"]

        final_fcst = apply_manual_adjustments(final_fcst, adj_dict=st.session_state.fcst_adj)

        # Write back into session as well
        R["final_fcst"] = final_fcst
        st.session_state.review_results = R

        tabs = st.tabs(["Final FCST", "Groups (M1)", "USW FCST (M2)", "ASHIP FCST (M3)", "Run Info"])

        # ── Final FCST tab
        with tabs[0]:
            st.markdown("### 📋 Final Table")

            # Utility to build the on-screen table (with overrides applied)
            def _build_final_display_table(final_fcst: pd.DataFrame, show_only_updates: bool) -> pd.DataFrame:
                # Apply overrides
                final_fcst = apply_update_needed_overrides(final_fcst)

                cols = [
                    "MATERIAL_KEY","Group","Segment","LongShort",
                    "Month","ds","source",
                    "FCST_final","previous_FCST",
                    "Selected_FCST_before_trend","trend_multiplier",
                    "SeasonalityFactor","BaselineUnits",
                    "USW_raw_per_week","Door_at_cutoff",
                    "FCST_Update_Needed","Update_Needed_eff",
                    "GrossChangePct","GrossChangePct_view","manual_adj_pct"
                ]
                show = [c for c in cols if c in final_fcst.columns]
                df = final_fcst[show].rename(columns={
                    "MATERIAL_KEY":"Material",
                    "Selected_FCST_before_trend":"Final_FCST_before_trend",
                    "SeasonalityFactor":"Seasonality",
                    "USW_raw_per_week":"U/S/W",
                    "Door_at_cutoff":"DoorCount",
                    "Update_Needed_eff":"Update Needed"
                }).copy()

                # === [PATCH-B] Add FR badge column =============================================
                if "FR_Needed" in final_fcst.columns:
                    # Merge FR_Needed by (Material, ds) and generate badge text
                    _fr_merge = (final_fcst[["MATERIAL_KEY","ds","FR_Needed"]]
                                    .rename(columns={"MATERIAL_KEY":"Material"}))
                    df = df.merge(_fr_merge, on=["Material","ds"], how="left")
                    df["FR_Needed"] = df["FR_Needed"].fillna(False).astype(bool)
                    df["FR"] = np.where(df["FR_Needed"], "⚠️ FR", "")
                else:
                    df["FR"] = ""
                # === [/PATCH-B] =============================================================


                if {"FCST_final","previous_FCST"}.issubset(df.columns):
                    numer = pd.to_numeric(df["FCST_final"], errors="coerce").astype(float)
                    denom = pd.to_numeric(df["previous_FCST"], errors="coerce").astype(float)
                    mask = (denom > 0).fillna(False)
                    diff_pct_val = np.full(len(df), np.nan, dtype=float)
                    diff_pct_val[mask] = (numer[mask] / denom[mask] - 1.0) * 100.0
                    df["FCST change %"] = pd.Series(diff_pct_val).map(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
                    df["FCST_previous"] = ["No FCST" if pd.isna(v) or v <= 0 else str(int(np.rint(v))) for v in denom]
                    df.drop(columns=["previous_FCST"], inplace=True, errors="ignore")

                for c in ["Final_FCST_before_trend", "FCST_final"]:
                    if c in df.columns:
                        df[c] = np.rint(pd.to_numeric(df[c], errors="coerce")).astype("Int64")
                if "manual_adj_pct" in df.columns:
                    df["Manual %"] = (
                        pd.to_numeric(df["manual_adj_pct"], errors="coerce") * 100
                    ).round(1).map(lambda x: f"{x:+.1f}%" if pd.notna(x) and x != 0 else "")
                    df.drop(columns=["manual_adj_pct"], inplace=True, errors="ignore")
                if "GrossChangePct_view" in df.columns:
                    df.rename(columns={"GrossChangePct_view":"FCST Gross %"}, inplace=True)

                desired_cols = [
                    "Material","Group","Segment","LongShort",
                    "Month","source",
                    "FCST_final","FCST_previous","FCST change %","FCST Gross %",
                    "Update Needed","FR",
                    "trend_multiplier",
                    "Seasonality","BaselineUnits","U/S/W","DoorCount",
                    "ds"
                ]
                df = df.reindex(columns=[c for c in desired_cols if c in df.columns])
                df = df.loc[:, ~df.columns.duplicated()].copy()

                if "GrossChangePct" in final_fcst.columns:
                    _abs_gross = pd.to_numeric(final_fcst["GrossChangePct"], errors="coerce").abs()
                    df["_abs_gross_"] = _abs_gross.values
                    if "Update Needed" in df.columns:
                        df.sort_values(["Update Needed","_abs_gross_"], ascending=[False, False], inplace=True)
                    df.drop(columns=["_abs_gross_"], inplace=True, errors="ignore")

                if show_only_updates and "Update Needed" in df.columns:
                    df = df[df["Update Needed"] == True]

                if "Update Needed" in df.columns:
                    df["Update Needed"] = df["Update Needed"].fillna(False).astype(bool)
                return df

            # Toggle
            show_only_updates = st.checkbox(
                "Show only items needing update (by Threshold (sum %))",
                value=False
            )

            display_df = _build_final_display_table(final_fcst, show_only_updates)

            # Editable table
            # ✅ Wrap in a form so changes apply only when 'Apply' is pressed
            with st.form("update_needed_apply_form", clear_on_submit=False):
                edited = st.data_editor(
                    display_df,
                    use_container_width=True,
                    height=520,
                    hide_index=True,
                    column_config={
                        "Material": st.column_config.TextColumn("Material", disabled=True),
                        "Group": st.column_config.TextColumn("Group", disabled=True),
                        "Segment": st.column_config.TextColumn("Segment", disabled=True),
                        "LongShort": st.column_config.TextColumn("Long/Short", disabled=True),
                        "Month": st.column_config.TextColumn("Month", disabled=True),
                        "source": st.column_config.TextColumn("source", disabled=True),
                        "FCST_final": st.column_config.NumberColumn("FCST_final", disabled=True),
                        "FCST_previous": st.column_config.TextColumn("FCST_previous", disabled=True),
                        "FCST change %": st.column_config.TextColumn("FCST change %", disabled=True),
                        "FCST Gross %": st.column_config.TextColumn("FCST Gross %", disabled=True),
                        # ⬇️ Checkboxes are editable (but not applied immediately)
                        "Update Needed": st.column_config.CheckboxColumn("Update Needed"),
                        "trend_multiplier": st.column_config.NumberColumn("trend_multiplier", disabled=True),
                        "Seasonality": st.column_config.NumberColumn("Seasonality", disabled=True),
                        "BaselineUnits": st.column_config.NumberColumn("BaselineUnits", disabled=True),
                        "U/S/W": st.column_config.NumberColumn("U/S/W", disabled=True),
                        "DoorCount": st.column_config.NumberColumn("DoorCount", disabled=True),
                        "ds": st.column_config.DatetimeColumn("ds", disabled=True),
                    },
                    key="final_fcst_editor",
                )

                # ⏩ Apply to session state only when the Apply button is pressed
                apply_updates = st.form_submit_button("✅ Apply (Update Needed)")

            if apply_updates:
                if not edited.empty and {"Update Needed", "ds", "Material"}.issubset(edited.columns):
                    new_overrides = {}
                    for _, r in edited.iterrows():
                        try:
                            k = (
                                str(r["Material"]),
                                pd.to_datetime(r["ds"]).to_period("M").to_timestamp(how="start"),
                            )
                            v = bool(r["Update Needed"])
                            new_overrides[k] = v
                        except Exception:
                            continue

                    # Actual apply point
                    st.session_state.update_needed_overrides.update(new_overrides)
                    st.success("Changes to Update Needed have been applied.")
                    st.rerun()


            # Export (.xlsx)
            st.markdown("### Export (.xlsx)")

            display_df_all = _build_final_display_table(final_fcst, show_only_updates=False)
            display_df_updates = display_df_all
            if "Update Needed" in display_df_updates.columns:
                display_df_updates = display_df_updates[display_df_updates["Update Needed"] == True]

            cA, cB, cC = st.columns(3)
            with cA:
                buf_whole = io.BytesIO()
                with pd.ExcelWriter(buf_whole, engine="openpyxl") as writer:
                    df_out = display_df_all.drop(columns=["ds"], errors="ignore")
                    df_out.to_excel(writer, index=False, sheet_name="Whole")
                st.download_button(
                    "⬇️ Download (Whole List)",
                    data=buf_whole.getvalue(),
                    file_name=f"FCST_Review_Final_{brand}_{cutoff_ts:%Y%m}_whole.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_whole_xlsx"
                )

            with cB:
                buf_updates = io.BytesIO()
                with pd.ExcelWriter(buf_updates, engine="openpyxl") as writer:
                    df_out = display_df_updates.drop(columns=["ds"], errors="ignore")
                    df_out.to_excel(writer, index=False, sheet_name="Updates")
                st.download_button(
                    "⬇️ Download (Update Needed only)",
                    data=buf_updates.getvalue(),
                    file_name=f"FCST_Review_Final_{brand}_{cutoff_ts:%Y%m}_updates.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_updates_xlsx"
                )

            with cC:
                # Report (Updated Needed only) — with overrides applied
                try:
                    final_fcst_for_report = apply_update_needed_overrides(st.session_state.review_results["final_fcst"])
                    # build_update_report_df uses FCST_Update_Needed, so
                    # For the report, temporarily replace FCST_Update_Needed to reflect overrides
                    tmp_report_df = final_fcst_for_report.copy()
                    tmp_report_df["FCST_Update_Needed"] = tmp_report_df["Update_Needed_eff"].fillna(False)
                    report_df = build_update_report_df(tmp_report_df, cutoff_ts)

                    from openpyxl import Workbook
                    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
                    from openpyxl.utils import get_column_letter

                    buf_report = io.BytesIO()
                    if report_df.empty:
                        wb = Workbook(); ws = wb.active; ws.title = "Report"
                        ws["A1"] = "No updated items after cutoff."
                        wb.save(buf_report)
                    else:
                        cols = list(report_df.columns)
                        old_cols = [c for c in cols if isinstance(c, tuple) and len(c) >= 2 and c[0] == "Old"]
                        new_cols = [c for c in cols if isinstance(c, tuple) and len(c) >= 2 and c[0] == "New"]

                        def _unique_keep_order(labels):
                            seen = set(); out = []
                            for x in labels:
                                lab = "" if x is None else str(x)
                                if lab not in seen:
                                    seen.add(lab); out.append(lab)
                            return out

                        old_month_labels = _unique_keep_order([c[1] for c in old_cols])
                        new_month_labels = _unique_keep_order([c[1] for c in new_cols])

                        wb = Workbook(); ws = wb.active; ws.title = "Report"
                        center = Alignment(horizontal="center", vertical="center")
                        bold = Font(bold=True)
                        th_fill = PatternFill("solid", fgColor="DDDDDD")
                        thin = Side(style="thin", color="AAAAAA")
                        border = Border(top=thin, left=thin, right=thin, bottom=thin)

                        ws.cell(row=1, column=1, value="ITEM")
                        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
                        ws["A1"].alignment = center; ws["A1"].font = bold; ws["A1"].fill = th_fill; ws["A1"].border = border

                        col_start_old = 2; col_end_old = col_start_old - 1
                        col_start_new = 2; col_end_new = col_start_new - 1

                        if old_month_labels:
                            col_end_old = col_start_old + len(old_month_labels) - 1
                            ws.cell(row=1, column=col_start_old, value="Old")
                            ws.merge_cells(start_row=1, start_column=col_start_old, end_row=1, end_column=col_end_old)
                            ws.cell(row=1, column=col_start_old).alignment = center
                            ws.cell(row=1, column=col_start_old).font = bold
                            for j, m in enumerate(old_month_labels, start=col_start_old):
                                ws.cell(row=2, column=j, value=str(m)).alignment = center
                                ws.cell(row=2, column=j).font = bold
                                ws.cell(row=1, column=j).fill = th_fill; ws.cell(row=2, column=j).fill = th_fill
                                ws.cell(row=1, column=j).border = border; ws.cell(row=2, column=j).border = border

                        col_start_new = (col_end_old + 1) if old_month_labels else 2
                        if new_month_labels:
                            col_end_new = col_start_new + len(new_month_labels) - 1
                            ws.cell(row=1, column=col_start_new, value="New")
                            ws.merge_cells(start_row=1, start_column=col_start_new, end_row=1, end_column=col_end_new)
                            ws.cell(row=1, column=col_start_new).alignment = center
                            ws.cell(row=1, column=col_start_new).font = bold
                            for j, m in enumerate(new_month_labels, start=col_start_new):
                                ws.cell(row=2, column=j, value=str(m)).alignment = center
                                ws.cell(row=2, column=j).font = bold
                                ws.cell(row=1, column=j).fill = th_fill; ws.cell(row=2, column=j).fill = th_fill
                                ws.cell(row=1, column=j).border = border; ws.cell(row=2, column=j).border = border

                        last_header_col = max(1, col_end_new, col_end_old)
                        old_pos = {lab: (col_start_old + i) for i, lab in enumerate(old_month_labels)} if old_month_labels else {}
                        new_pos = {lab: (col_start_new + i) for i, lab in enumerate(new_month_labels)} if new_month_labels else {}
                        old_key_by_label = {}
                        for c in old_cols:
                            lab = str(c[1])
                            if lab not in old_key_by_label: old_key_by_label[lab] = c
                        new_key_by_label = {}
                        for c in new_cols:
                            lab = str(c[1])
                            if lab not in new_key_by_label: new_key_by_label[lab] = c

                        import numpy as np
                        row_idx = 3
                        for i in range(len(report_df)):
                            item_val = report_df["ITEM"].iat[i] if "ITEM" in report_df.columns and i < len(report_df) else None
                            ws.cell(row=row_idx, column=1, value=("" if pd.isna(item_val) else str(item_val)))
                            ws.cell(row=row_idx, column=1).border = border

                            for m in old_month_labels:
                                col_idx = old_pos.get(m); 
                                if col_idx is None: continue
                                col_key = old_key_by_label.get(m, None)
                                v = np.nan if col_key is None else report_df.loc[i, col_key]
                                try:
                                    cell_value = None if pd.isna(v) else int(np.rint(float(v)))
                                except Exception:
                                    cell_value = None
                                ws.cell(row=row_idx, column=col_idx, value=cell_value).border = border

                            for m in new_month_labels:
                                col_idx = new_pos.get(m);
                                if col_idx is None: continue
                                col_key = new_key_by_label.get(m, None)
                                v = np.nan if col_key is None else report_df.loc[i, col_key]
                                try:
                                    cell_value = None if pd.isna(v) else int(np.rint(float(v)))
                                except Exception:
                                    cell_value = None
                                ws.cell(row=row_idx, column=col_idx, value=cell_value).border = border

                            row_idx += 1

                        ws.column_dimensions["A"].width = 18
                        for j in range(2, last_header_col + 1):
                            ws.column_dimensions[get_column_letter(j)].width = 12
                        for r in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=last_header_col):
                            for cell in r:
                                cell.alignment = center
                                if not cell.font.bold: cell.font = bold
                                cell.border = border
                                if cell.fill.fill_type is None: cell.fill = th_fill

                        wb.save(buf_report)

                    st.download_button(
                        label="📥 Download Report Format",
                        data=buf_report.getvalue(),
                        file_name=f"FCST_Report_UpdatedNeeded_{pd.Period(cutoff_ts, 'M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_report_xlsx"
                    )
                except Exception as e:
                    st.warning(f"Report 생성 중 오류: {e}")

        # ── Groups (M1)
        with tabs[1]:
            st.dataframe(review_df, use_container_width=True, height=420)
            df_to_excel_download(
                review_df,
                f"FCST_Review_Groups_{brand}_{cutoff_ts:%Y%m}.xlsx",
                "⬇️ Groups Download"
            )

        # ── USW FCST (M2)
        with tabs[2]:
            st.dataframe(usw_fcst, use_container_width=True, height=420)
            df_to_excel_download(
                usw_fcst,
                f"USW_FCST_{brand}_{cutoff_ts:%Y%m}.xlsx",
                "⬇️ USW Download"
            )

        # ── ASHIP FCST (M3)
        with tabs[3]:
            st.dataframe(aship_fcst, use_container_width=True, height=420)
            df_to_excel_download(
                aship_fcst,
                f"ASHIP_FCST_{brand}_{cutoff_ts:%Y%m}.xlsx",
                "⬇️ ASHIP Download"
            )

        # ── Run Info
        with tabs[4]:
            st.json({"USW": usw_info, "ASHIP": aship_info})
            st.caption("Trend settings are fixed in code and applied at the final step (FCST × trend_multiplier).")

        render_global_graph(st.session_state.review_results, cutoff_ts)

    except Exception as e:
        st.exception(e)
